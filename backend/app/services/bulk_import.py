import csv
import io
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models.product import Category, Product, ProductVariant

EXPECTED_COLUMNS = [
    "product_name",
    "brand",
    "category",
    "hsn_code",
    "tax_rate",
    "sku",
    "barcode",
    "size",
    "color",
    "cost_price",
    "selling_price",
    "mrp",
    "reorder_level",
]


def parse_rows(filename: str, contents: bytes) -> list[dict]:
    lowered = filename.lower()
    if lowered.endswith(".xlsx"):
        return _parse_xlsx(contents)
    if lowered.endswith(".csv"):
        return _parse_csv(contents)
    raise ValueError("Only .csv or .xlsx files are supported")


def _parse_csv(contents: bytes) -> list[dict]:
    text = contents.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        cleaned = {(k or "").strip().lower(): (v or "").strip() for k, v in row.items() if k}
        if any(cleaned.values()):
            rows.append(cleaned)
    return rows


def _parse_xlsx(contents: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
    rows = []
    for values in rows_iter:
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        row = {headers[i]: ("" if v is None else str(v).strip()) for i, v in enumerate(values) if i < len(headers)}
        rows.append(row)
    return rows


def _to_float(value: str | None, default: float) -> float:
    if not value:
        return default
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, ValueError):
        return default


def _to_int(value: str | None, default: int) -> int:
    if not value:
        return default
    try:
        return int(float(value))
    except ValueError:
        return default


def import_rows(db: Session, rows: list[dict]) -> dict:
    created_products = 0
    updated_products = 0
    created_variants = 0
    updated_variants = 0
    errors: list[dict] = []
    product_cache: dict[tuple[str, str], Product] = {}

    for idx, row in enumerate(rows, start=2):
        product_name = (row.get("product_name") or "").strip()
        sku = (row.get("sku") or "").strip()
        if not product_name or not sku:
            errors.append({"row": idx, "message": "product_name and sku are required"})
            continue

        brand = (row.get("brand") or "").strip() or None
        category_name = (row.get("category") or "").strip() or None

        category = None
        if category_name:
            category = db.query(Category).filter(Category.name.ilike(category_name)).first()
            if not category:
                category = Category(name=category_name)
                db.add(category)
                db.flush()

        cache_key = (product_name.lower(), (brand or "").lower())
        product = product_cache.get(cache_key)
        if not product:
            query = db.query(Product).filter(Product.name.ilike(product_name))
            query = query.filter(Product.brand.ilike(brand)) if brand else query.filter(Product.brand.is_(None))
            product = query.first()

        if not product:
            product = Product(
                name=product_name,
                brand=brand,
                category_id=category.id if category else None,
                hsn_code=(row.get("hsn_code") or "").strip() or None,
                tax_rate=_to_float(row.get("tax_rate"), 0),
            )
            db.add(product)
            db.flush()
            created_products += 1
        else:
            changed = False
            if category and product.category_id != category.id:
                product.category_id = category.id
                changed = True
            if row.get("hsn_code") and row["hsn_code"] != product.hsn_code:
                product.hsn_code = row["hsn_code"].strip()
                changed = True
            if row.get("tax_rate"):
                new_rate = _to_float(row.get("tax_rate"), float(product.tax_rate))
                if new_rate != float(product.tax_rate):
                    product.tax_rate = new_rate
                    changed = True
            if changed:
                updated_products += 1

        product_cache[cache_key] = product

        try:
            variant_data = dict(
                barcode=(row.get("barcode") or "").strip() or None,
                size=(row.get("size") or "").strip() or None,
                color=(row.get("color") or "").strip() or None,
                cost_price=_to_float(row.get("cost_price"), 0),
                selling_price=_to_float(row.get("selling_price"), 0),
                mrp=_to_float(row.get("mrp"), 0),
                reorder_level=_to_int(row.get("reorder_level"), 5),
            )
        except Exception as exc:  # defensive: malformed numeric cell
            errors.append({"row": idx, "message": f"Invalid value: {exc}"})
            continue

        variant = db.query(ProductVariant).filter(ProductVariant.sku == sku).first()
        if variant:
            for field, value in variant_data.items():
                setattr(variant, field, value)
            updated_variants += 1
        else:
            variant = ProductVariant(product_id=product.id, sku=sku, **variant_data)
            db.add(variant)
            db.flush()
            created_variants += 1

    return {
        "created_products": created_products,
        "updated_products": updated_products,
        "created_variants": created_variants,
        "updated_variants": updated_variants,
        "total_rows": len(rows),
        "errors": errors,
    }


def build_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(EXPECTED_COLUMNS)
    ws.append(
        [
            "Banarasi Silk Saree",
            "Tanisi",
            "Sarees",
            "5407",
            "5",
            "BSS-RED-FS",
            "8901000000099",
            "Free Size",
            "Red",
            "1800",
            "3200",
            "3500",
            "5",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
