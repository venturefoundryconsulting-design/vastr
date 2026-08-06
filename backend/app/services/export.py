"""Reusable export engine - turns a list of flat dicts into a downloadable file.

Every module (Products, Sales, Customers, Vendors, Purchase Orders, Inventory,
Reports) calls `build_export_response()` with the same shape: rows already
flattened to plain dicts, and a list of (field_key, header_label) columns. Adding
export to a new module never touches this file - just flatten your existing list
data and call it, same as every other module already does.
"""

import csv
import io
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from typing import Any, Literal

from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

ExportFormat = Literal["csv", "xlsx", "pdf"]
Column = tuple[str, str]  # (row dict key, header label)

BRAND = colors.HexColor("#9d174d")


def _cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def build_export_response(
    rows: list[dict], columns: list[Column], title: str, fmt: ExportFormat
) -> Response:
    filename_base = title.lower().replace(" ", "_")
    if fmt == "csv":
        return _to_csv(rows, columns, filename_base)
    if fmt == "xlsx":
        return _to_xlsx(rows, columns, title, filename_base)
    if fmt == "pdf":
        return _to_pdf(rows, columns, title, filename_base)
    raise ValueError(f"Unsupported export format: {fmt}")


def _to_csv(rows: list[dict], columns: list[Column], filename_base: str) -> Response:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for _, label in columns])
    for row in rows:
        writer.writerow([_cell(row.get(key)) for key, _ in columns])
    return Response(
        content=buf.getvalue().encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
    )


def _to_xlsx(rows: list[dict], columns: list[Column], title: str, filename_base: str) -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Export"
    header_font = Font(bold=True)
    for col_idx, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell(row.get(key)))
    for col_idx, (key, label) in enumerate(columns, start=1):
        max_len = max([len(label)] + [len(str(_cell(r.get(key)))) for r in rows[:200]], default=10)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
    )


def _to_pdf(rows: list[dict], columns: list[Column], title: str, filename_base: str) -> Response:
    buf = io.BytesIO()
    landscape_needed = len(columns) > 5
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4) if landscape_needed else A4,
        topMargin=18 * mm,
        bottomMargin=14 * mm,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ExportTitle", parent=styles["Heading1"], textColor=BRAND, spaceAfter=4)
    meta_style = ParagraphStyle("ExportMeta", parent=styles["Normal"], textColor=colors.grey, fontSize=9)

    elements = [
        Paragraph(title, title_style),
        Paragraph(
            f"Generated {datetime.now().strftime('%d %b %Y, %I:%M %p')} · {len(rows)} row(s)", meta_style
        ),
        Spacer(1, 10),
    ]

    header = [label for _, label in columns]
    data = [header] + [[str(_cell(row.get(key))) for key, _ in columns] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), BRAND),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5d9df")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#faf7f8")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)

    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'},
    )
